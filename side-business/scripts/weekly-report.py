#!/usr/bin/env python3
"""記録簿.xlsx から週次レポートを生成する。

使い方:
    python3 scripts/weekly-report.py                 # 今週分を logs/ に書き出す
    python3 scripts/weekly-report.py --print         # 標準出力に出すだけ
    python3 scripts/weekly-report.py --week 2026-W40 # 週を指定

数式のキャッシュ値には依存せず、「記録」シートの生データから直接集計する。
"""
import argparse, datetime as dt, pathlib, sys
from collections import defaultdict

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("openpyxl が入っていません: pip install openpyxl")

ROOT = pathlib.Path(__file__).resolve().parent.parent
LEDGER = ROOT / "記録簿.xlsx"
LOGS = ROOT / "logs"
START = dt.date(2026, 9, 1)          # プロジェクト開始日。実際の着手日に合わせて変更する
TARGET_M1 = 1000                     # M1: 月商1,000円


def load_entries():
    """記録シートを (日付, 種別, 内容, 金額, プラットフォーム) のリストで返す。"""
    if not LEDGER.exists():
        sys.exit(f"{LEDGER} が見つかりません")
    ws = load_workbook(LEDGER, data_only=True)["記録"]
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        date, kind, desc, amount, platform = row[0], row[1], row[2], row[3], row[4]
        if date is None or kind not in ("売上", "経費") or amount is None:
            continue
        if isinstance(date, dt.datetime):
            date = date.date()
        out.append((date, kind, desc or "", float(amount), platform or ""))
    return out


def week_bounds(label=None):
    """ISO週のラベルから月曜〜日曜を返す。省略時は今週。"""
    if label:
        y, w = label.upper().split("-W")
        monday = dt.date.fromisocalendar(int(y), int(w), 1)
    else:
        today = dt.date.today()
        monday = today - dt.timedelta(days=today.weekday())
    return monday, monday + dt.timedelta(days=6)


def total(entries, kind, lo, hi):
    return sum(a for d, k, _, a, _ in entries if k == kind and lo <= d <= hi)


def count_sales(entries, lo, hi):
    return sum(1 for d, k, _, _, _ in entries if k == "売上" and lo <= d <= hi)


def build(entries, monday, sunday):
    ymd = "%Y-%m-%d"
    m_lo = sunday.replace(day=1)
    m_hi = sunday
    prev_lo = monday - dt.timedelta(days=7)
    prev_hi = monday - dt.timedelta(days=1)

    w_sales = total(entries, "売上", monday, sunday)
    w_cost = total(entries, "経費", monday, sunday)
    w_n = count_sales(entries, monday, sunday)
    p_sales = total(entries, "売上", prev_lo, prev_hi)
    p_n = count_sales(entries, prev_lo, prev_hi)
    m_sales = total(entries, "売上", m_lo, m_hi)
    m_cost = total(entries, "経費", m_lo, m_hi)
    m_n = count_sales(entries, m_lo, m_hi)
    all_sales = total(entries, "売上", dt.date.min, sunday)
    all_cost = total(entries, "経費", dt.date.min, sunday)
    all_n = count_sales(entries, dt.date.min, sunday)

    day = (sunday - START).days + 1
    phase = ("Phase 1（〜Day 30）" if day <= 30 else
             "Phase 2（Day 31〜60）" if day <= 60 else
             "Phase 3（Day 61〜90）" if day <= 90 else "Day 90 以降")
    pct = int(m_sales / TARGET_M1 * 100) if TARGET_M1 else 0
    bar = "█" * min(20, pct // 5) + "░" * (20 - min(20, pct // 5))

    def d(cur, prev):
        if prev == 0:
            return "—" if cur == 0 else "新規"
        return f"{cur - prev:+,.0f}円"

    # プラットフォーム別
    by_pf = defaultdict(float)
    for dd, k, _, a, pf in entries:
        if k == "売上" and m_lo <= dd <= m_hi:
            by_pf[pf or "(未記入)"] += a
    pf_rows = "\n".join(f"| {k} | ¥{v:,.0f} |" for k, v in sorted(by_pf.items(), key=lambda x: -x[1])) \
              or "| （売上なし） | — |"

    L = [
        f"# 週次レポート {monday.isocalendar().year}-W{monday.isocalendar().week:02d}",
        "",
        f"期間: {monday:{ymd}} 〜 {sunday:{ymd}} / Day {day} / {phase}",
        "",
        "## 3指標",
        "",
        "| 指標 | 今週 | 先週比 | 今月 | 累計 |",
        "|---|---|---|---|---|",
        f"| 売上 | ¥{w_sales:,.0f} | {d(w_sales, p_sales)} | ¥{m_sales:,.0f} | ¥{all_sales:,.0f} |",
        f"| **購入件数** | {w_n} 件 | {w_n - p_n:+d} 件 | {m_n} 件 | {all_n} 件 |",
        f"| 経費 | ¥{w_cost:,.0f} | — | ¥{m_cost:,.0f} | ¥{all_cost:,.0f} |",
        f"| 所得 | ¥{w_sales - w_cost:,.0f} | — | ¥{m_sales - m_cost:,.0f} | ¥{all_sales - all_cost:,.0f} |",
        "",
        "## M1（月商1,000円）への進捗",
        "",
        f"```\n{bar} {pct}%   ¥{m_sales:,.0f} / ¥{TARGET_M1:,}\n```",
        "",
        "## 今月の売上（プラットフォーム別）",
        "",
        "| プラットフォーム | 売上 |",
        "|---|---|",
        pf_rows,
        "",
        "---",
        "",
        "## ここから下は手で書く（5分）",
        "",
        "### 公開したもの",
        "- ",
        "",
        "### 効いた / 効かなかった",
        "- 効いた: ",
        "- 効かなかった: ",
        "",
        "### 来週やること（3つまで）",
        "1. ",
        "2. ",
        "3. ",
        "",
        "### 詰まっていること",
        "- ",
        "",
        "### 使った時間",
        "",
        "| | 時間 |",
        "|---|---|",
        "| 執筆・実装 | |",
        "| 告知・アウトリーチ | |",
    ]

    # 判定と警告
    warn = []
    if day >= 30 and all_n == 0 and m_sales == 0:
        warn.append("Day 30 を過ぎて売上ゼロ。**商品ではなく告知量を先に疑うこと**（11-90day-plan.md の Day 60 判定）")
    if day >= 60 and all_sales == 0:
        warn.append("Day 60 を過ぎて売上ゼロ。記事末尾に商品リンクがあるか、X で告知したのが何回かを確認する")
    if day >= 90 and m_sales < TARGET_M1:
        warn.append("Day 90 で M1 未達。撤退ではなく、受託を1件発動する（10-strategy.md §4）")
    if all_cost > 0 and all_sales == 0 and day >= 90:
        warn.append("3ヶ月間、支出のみで売上ゼロ。固定費を点検する")
    if warn:
        L += ["", "---", "", "## ⚠ 判定"] + [f"- {w}" for w in warn]

    L += ["", "---", "", f"*生成: {dt.date.today():{ymd}} / `scripts/weekly-report.py`*"]
    return "\n".join(L)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--week", help="ISO週 (例 2026-W40)。省略時は今週")
    ap.add_argument("--print", action="store_true", dest="stdout", help="ファイルに書かず標準出力へ")
    args = ap.parse_args()

    monday, sunday = week_bounds(args.week)
    report = build(load_entries(), monday, sunday)

    if args.stdout:
        print(report)
        return
    LOGS.mkdir(exist_ok=True)
    path = LOGS / f"{monday.isocalendar().year}-W{monday.isocalendar().week:02d}.md"
    if path.exists():
        print(f"既に存在します: {path}\n上書きしたくない場合は --print を使ってください")
        return
    path.write_text(report, encoding="utf-8")
    print(f"書き出しました: {path}")


if __name__ == "__main__":
    main()
